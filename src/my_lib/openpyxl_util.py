#!/usr/bin/env python3
from __future__ import annotations

import logging
import pathlib
from typing import Any, Protocol

import openpyxl.drawing.image
import openpyxl.drawing.spreadsheet_drawing
import openpyxl.styles
import openpyxl.utils
import openpyxl.utils.units
import openpyxl.workbook
import openpyxl.worksheet.worksheet


class ThumbPathFunc(Protocol):
    """サムネイル画像のパスを取得するコールバック関数の型"""

    def __call__(self, item: dict[str, Any]) -> pathlib.Path | None:
        """アイテムからサムネイル画像のパスを取得する。

        Args:
            item: 商品情報を含む辞書

        Returns:
            サムネイル画像のパス。存在しない場合は None
        """
        ...


class SetStatusFunc(Protocol):
    """ステータス表示を更新するコールバック関数の型"""

    def __call__(self, status: str) -> None:
        """ステータスメッセージを設定する。

        Args:
            status: 表示するステータスメッセージ
        """
        ...


class UpdateFunc(Protocol):
    """進捗を更新するコールバック関数の型"""

    def __call__(self) -> None:
        """進捗を1つ進める。"""
        ...


def gen_text_pos(row: int, col: int) -> str:
    return f"{openpyxl.utils.get_column_letter(col)}{row}"


def set_header_cell_style(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    value: str,
    width: float | None,
    style: dict[str, Any],
) -> None:  # noqa: PLR0913
    sheet.cell(row, col).value = value
    sheet.cell(row, col).style = "Normal"
    sheet.cell(row, col).border = style["border"]
    sheet.cell(row, col).fill = style["fill"]

    if width is not None:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width


def insert_table_header(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    sheet_def: dict[str, Any],
    base_style: dict[str, Any],
) -> None:
    for key in sheet_def["TABLE_HEADER"]["col"]:
        col: int = sheet_def["TABLE_HEADER"]["col"][key]["pos"]
        width: float | None = sheet_def["TABLE_HEADER"]["col"][key].get("width", None)

        if key == "category":
            for i in range(sheet_def["TABLE_HEADER"]["col"][key]["length"]):
                set_header_cell_style(
                    sheet,
                    row,
                    col + i,
                    sheet_def["TABLE_HEADER"]["col"][key]["label"] + f" ({i + 1})",
                    width,
                    base_style,
                )
        else:
            set_header_cell_style(
                sheet, row, col, sheet_def["TABLE_HEADER"]["col"][key]["label"], width, base_style
            )


def gen_item_cell_style(base_style: dict[str, Any], cell_def: dict[str, Any]) -> dict[str, Any]:
    style = base_style.copy()

    if "format" in cell_def:
        style["text_format"] = cell_def["format"]

    style["text_wrap"] = cell_def.get("wrap", False)

    return style


def set_item_cell_style(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    value: Any,
    style: dict[str, Any],
) -> None:
    sheet.cell(row, col).value = value
    sheet.cell(row, col).style = "Normal"
    sheet.cell(row, col).border = style["border"]
    sheet.cell(row, col).alignment = openpyxl.styles.Alignment(wrap_text=style["text_wrap"], vertical="top")

    if "text_format" in style:
        sheet.cell(row, col).number_format = style["text_format"]


def insert_table_item(  # noqa: PLR0912, PLR0913, C901
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    item: dict[str, Any],
    is_need_thumb: bool,
    thumb_path: pathlib.Path | None,
    sheet_def: dict[str, Any],
    base_style: dict[str, Any],
) -> None:
    for key in sheet_def["TABLE_HEADER"]["col"]:
        col: int = sheet_def["TABLE_HEADER"]["col"][key]["pos"]

        cell_style = gen_item_cell_style(base_style, sheet_def["TABLE_HEADER"]["col"][key])

        if key == "category":
            if key in item:
                for i in range(sheet_def["TABLE_HEADER"]["col"][key]["length"]):
                    value: Any = item[key][i] if i < len(item["category"]) else ""
                    set_item_cell_style(sheet, row, col + i, value, cell_style)
            else:
                logging.warning("キー '%s' がアイテムに存在しません", key)
                value = None
        elif key == "image":
            sheet.cell(row, col).border = cell_style["border"]
            if is_need_thumb:
                insert_table_cell_image(
                    sheet,
                    row,
                    col,
                    thumb_path,
                    sheet_def["TABLE_HEADER"]["col"]["image"]["width"],
                    sheet_def["TABLE_HEADER"]["row"]["height"]["default"],
                )
        else:
            if (
                ("optional" in sheet_def["TABLE_HEADER"]["col"][key])
                and sheet_def["TABLE_HEADER"]["col"][key]["optional"]
                and (key not in item)
            ):
                value = None
            else:
                if "value" in sheet_def["TABLE_HEADER"]["col"][key]:
                    value = sheet_def["TABLE_HEADER"]["col"][key]["value"]
                elif "formal_key" in sheet_def["TABLE_HEADER"]["col"][key]:
                    if sheet_def["TABLE_HEADER"]["col"][key]["formal_key"] in item:
                        value = item[sheet_def["TABLE_HEADER"]["col"][key]["formal_key"]]
                    else:
                        logging.warning(
                            "キー '%s' (formal_key: '%s') がアイテムに存在しません",
                            key,
                            sheet_def["TABLE_HEADER"]["col"][key]["formal_key"],
                        )
                        value = None
                else:
                    if key in item:
                        value = item[key]
                    else:
                        logging.warning("キー '%s' がアイテムに存在しません", key)
                        value = None

                if "conv_func" in sheet_def["TABLE_HEADER"]["col"][key]:
                    value = sheet_def["TABLE_HEADER"]["col"][key]["conv_func"](value)

            set_item_cell_style(sheet, row, col, value, cell_style)

        if "link_func" in sheet_def["TABLE_HEADER"]["col"][key]:
            sheet.cell(row, col).hyperlink = sheet_def["TABLE_HEADER"]["col"][key]["link_func"](item)


def insert_table_cell_image(  # noqa: PLR0913
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    thumb_path: pathlib.Path | None,
    cell_width: float,
    cell_height: float,
) -> None:
    if (thumb_path is None) or (not thumb_path.exists()):
        return

    img = openpyxl.drawing.image.Image(thumb_path)

    # NOTE: マジックナンバー「8」は下記等を参考にして設定。(日本語フォントだと 8 が良さそう)
    # > In all honesty, I cannot tell you how many blogs and stack overflow answers
    # > I read before I stumbled across this magic number: 7.5
    # https://imranhugo.medium.com/how-to-right-align-an-image-in-excel-cell-using-python-and-openpyxl-7ca75a85b13a
    cell_width_pix = cell_width * 8
    cell_height_pix = openpyxl.utils.units.points_to_pixels(cell_height)

    cell_width_emu = openpyxl.utils.units.pixels_to_EMU(cell_width_pix)
    cell_height_emu = openpyxl.utils.units.pixels_to_EMU(cell_height_pix)

    margin_pix = 2
    content_width_pix = cell_width_pix - (margin_pix * 2)
    content_height_pix = cell_height_pix - (margin_pix * 2)

    content_ratio = content_width_pix / content_height_pix
    image_ratio = img.width / img.height

    if (img.width > content_width_pix) or (img.height > content_height_pix):
        if image_ratio > content_ratio:
            # NOTE: 画像の横幅をセルの横幅に合わせる
            scale = content_width_pix / img.width
        else:
            # NOTE: 画像の高さをセルの高さに合わせる
            scale = content_height_pix / img.height

        img.width *= scale
        img.height *= scale

    image_width_emu = openpyxl.utils.units.pixels_to_EMU(img.width)
    image_height_emu = openpyxl.utils.units.pixels_to_EMU(img.height)

    col_offset_emu = (cell_width_emu - image_width_emu) / 2
    row_offset_emu = (cell_height_emu - image_height_emu) / 2

    marker_1 = openpyxl.drawing.spreadsheet_drawing.AnchorMarker(
        col=col - 1, row=row - 1, colOff=col_offset_emu, rowOff=row_offset_emu
    )
    marker_2 = openpyxl.drawing.spreadsheet_drawing.AnchorMarker(
        col=col, row=row, colOff=-col_offset_emu, rowOff=-row_offset_emu
    )

    img.anchor = openpyxl.drawing.spreadsheet_drawing.TwoCellAnchor(_from=marker_1, to=marker_2)

    sheet.add_image(img)


def setting_table_view(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    sheet_def: dict[str, Any],
    row_last: int,
    is_hidden: bool,
) -> None:
    sheet.column_dimensions.group(
        openpyxl.utils.get_column_letter(sheet_def["TABLE_HEADER"]["col"]["image"]["pos"]),
        openpyxl.utils.get_column_letter(sheet_def["TABLE_HEADER"]["col"]["image"]["pos"]),
        hidden=is_hidden,
    )

    sheet.freeze_panes = gen_text_pos(
        sheet_def["TABLE_HEADER"]["row"]["pos"] + 1,
        sheet_def["TABLE_HEADER"]["col"]["price"]["pos"] + 1,
    )

    sheet.auto_filter.ref = "{start}:{end}".format(
        start=gen_text_pos(
            sheet_def["TABLE_HEADER"]["row"]["pos"],
            min([x["pos"] for x in sheet_def["TABLE_HEADER"]["col"].values()]),
        ),
        end=gen_text_pos(row_last, max([x["pos"] for x in sheet_def["TABLE_HEADER"]["col"].values()])),
    )
    sheet.sheet_view.showGridLines = False


def generate_list_sheet(  # noqa: PLR0913
    book: openpyxl.workbook.Workbook,
    item_list: list[dict[str, Any]],
    sheet_def: dict[str, Any],
    is_need_thumb: bool,
    thumb_path_func: ThumbPathFunc,
    set_status_func: SetStatusFunc,
    update_seq_func: UpdateFunc,
    update_item_func: UpdateFunc,
) -> openpyxl.worksheet.worksheet.Worksheet:
    sheet = book.create_sheet()
    sheet.title = "{label}アイテム一覧".format(label=sheet_def["SHEET_TITLE"])

    side = openpyxl.styles.Side(border_style="thin", color="000000")
    border = openpyxl.styles.Border(top=side, left=side, right=side, bottom=side)
    fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="F2F2F2")

    base_style: dict[str, Any] = {"border": border, "fill": fill}

    row = sheet_def["TABLE_HEADER"]["row"]["pos"]

    set_status_func("テーブルのヘッダを設定しています...")
    insert_table_header(sheet, row, sheet_def, base_style)

    update_seq_func()

    set_status_func("{label} - 商品の記載をしています...".format(label=sheet_def["SHEET_TITLE"]))

    cell_height: float
    if is_need_thumb:
        cell_height = sheet_def["TABLE_HEADER"]["row"]["height"]["default"]
    else:
        cell_height = sheet_def["TABLE_HEADER"]["row"]["height"]["without_thumb"]

    row += 1
    for item in item_list:
        sheet.row_dimensions[row].height = cell_height
        insert_table_item(sheet, row, item, is_need_thumb, thumb_path_func(item), sheet_def, base_style)
        update_item_func()

        row += 1

    row_last = row - 1

    update_item_func()
    update_seq_func()

    set_status_func("テーブルの表示設定しています...")
    setting_table_view(sheet, sheet_def, row_last, not is_need_thumb)

    update_seq_func()

    return sheet
